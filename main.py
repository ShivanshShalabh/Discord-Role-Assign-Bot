from asyncio.coroutines import coroutine
from io import StringIO
from time import sleep
from discord.ext import commands
from discord.utils import get
import discord
import openpyxl
import requests
client = commands.Bot(command_prefix='!', help_command=None)
workbook = openpyxl.load_workbook('Tech Club.xlsx')
sheet = workbook.active


header = {
    'authorization': 'auth'
}


def sendMsg(cnt):
    payload = {
        'content': cnt
    }
    r = requests.post(
        "url", data=payload, headers=header)


def findMsg(fname, sname, _class, sec, author):
    role = ""
    check = True
    for i in range(2, 195):
        names = sheet.cell(row=i, column=2).value
        namelst = names.split(" ")
        class_ = sheet.cell(row=i, column=3).value
        sec_ = sheet.cell(row=i, column=4).value
        if(namelst[0].lower() == fname.lower() and namelst[1].lower() == sname.lower() and str(_class) == str(class_) and sec_ == sec):
            if sheet.cell(row=i, column=12).value == -1:
                role = sheet.cell(row=i, column=5).value
                role_lst = role.split(", ")
                sheet.cell(row=i, column=12).value = author
                ret_lst = [names]
                workbook.save('Tech Club.xlsx')
            else:
                print(sheet.cell(row=i, column=12).value)
                check = False
            break
    if(role == "" and check == True):
        print("Invalid request! If you think this is an error, please contact an admin.")
        return "Invalid request! Make sure that the details you are entering are right and try again.\nIf you think this is an error, please contact an admin."
    elif (check == False):
        print("Duplicate request")
        return "Duplicate request! This student is already registered. If you think this is an error, please contact an admin."
    else:
        if "Competitive programming" in role_lst:
            ret_lst.append('Test-CP')

        if "Development" in role_lst:
            ret_lst.append('Test-Development')

        if "Robotics" in role_lst:
            ret_lst.append('Test-Robotics')
        print(f"Found {ret_lst}")
        return ret_lst


@client.event
async def on_ready():
    print(f'{client.user} has Awoken!')


@client.command()
async def verify(ctx, fname, sname, class_, sec):
    print(f"Message Accepted: !verify {fname} {sname} {class_} { sec}")
    comands = []
    data_lst = findMsg(fname, sname, class_, sec, str(
        f'{ctx.message.author.name}, {ctx.message.author.id}'))
    if type(data_lst) == str:
        sendMsg(data_lst)
    else:
        for k in range(1, len(data_lst)):
            msg = ((f'!addroles Applicants {ctx.author.mention}'))
            sendMsg(msg)


@client.command()
@commands.has_any_role(...["{LIST OF ROLES}"])
async def addroles(ctx, role: discord.Role, user: discord.Member):
    await user.add_roles(role)
    await ctx.send(f"Successfully assigned the role {role} to {user}")


client.run('key')
